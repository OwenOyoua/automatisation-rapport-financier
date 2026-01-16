"""
Exporteur de rapports au format Excel (XLSX).
Génère des classeurs Excel avec formules et graphiques.
"""

import os
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.models import BilanFonctionnel, BilanFinancier, PatrimoineEntreprise


class ExcelExporter:
    """
    Exporteur pour générer des rapports Excel professionnels.
    """

    def __init__(self):
        self.setup_styles()

    def setup_styles(self):
        """Configurer les styles pour Excel."""
        # Style pour les titres
        self.title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        self.title_alignment = Alignment(horizontal='center', vertical='center')
        
        # Style pour les sous-titres
        self.subtitle_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.subtitle_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # Style pour les en-têtes de tableau
        self.header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Style pour les totaux
        self.total_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        self.total_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        
        # Style pour les nombres
        self.number_alignment = Alignment(horizontal='right')
        
        # Bordures
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export(self, report_data, filename: str, options: Dict[str, Any]) -> str:
        """
        Exporter un rapport au format Excel.
        
        Args:
            report_data: Données du rapport (BilanFonctionnel, BilanFinancier, etc.)
            filename: Nom du fichier de sortie
            options: Options d'exportation
            
        Returns:
            Chemin du fichier généré
        """
        # Créer le répertoire d'export si nécessaire
        export_dir = Path("exports")
        export_dir.mkdir(exist_ok=True)
        
        file_path = export_dir / filename
        
        # Créer le classeur
        wb = Workbook()
        
        # Supprimer la feuille par défaut
        wb.remove(wb.active)
        
        # Contenu du classeur
        if isinstance(report_data, BilanFonctionnel):
            self.create_bilan_fonctionnel_sheets(wb, report_data, options)
        elif isinstance(report_data, BilanFinancier):
            self.create_bilan_financier_sheets(wb, report_data, options)
        elif isinstance(report_data, PatrimoineEntreprise):
            self.create_patrimoine_sheets(wb, report_data, options)
        
        # Feuille de résumé
        self.create_summary_sheet(wb, report_data, options)
        
        # Sauvegarder le classeur
        wb.save(str(file_path))
        
        return str(file_path)

    def create_bilan_fonctionnel_sheets(self, wb: Workbook, bilan: BilanFonctionnel, options: Dict[str, Any]):
        """Créer les feuilles pour le bilan fonctionnel."""
        # Feuille principale
        ws = wb.create_sheet("Bilan Fonctionnel")
        self.create_bilan_fonctionnel_main(ws, bilan, options)
        
        # Feuille d'analyse
        ws_analysis = wb.create_sheet("Analyse")
        self.create_bilan_fonctionnel_analysis(ws_analysis, bilan, options)
        
        # Feuille de graphiques
        if options.get('include_charts', True):
            ws_charts = wb.create_sheet("Graphiques")
            self.create_bilan_fonctionnel_charts(ws_charts, bilan, options)

    def create_bilan_fonctionnel_main(self, ws: Worksheet, bilan: BilanFonctionnel, options: Dict[str, Any]):
        """Créer la feuille principale du bilan fonctionnel."""
        # Titre
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = "BILAN FONCTIONNEL"
        title_cell.font = self.title_font
        title_cell.fill = self.title_fill
        title_cell.alignment = self.title_alignment
        
        # Informations
        ws['A3'] = "Entreprise:"
        ws['B3'] = options.get('entreprise', 'Entreprise')
        ws['A4'] = "Période:"
        ws['B4'] = options.get('periode', '2024')
        ws['A5'] = "Devise:"
        ws['B5'] = options.get('devise', 'MAD')
        ws['A6'] = "Date:"
        ws['B6'] = datetime.now().strftime('%d/%m/%Y')
        
        # Tableau principal
        row_start = 8
        
        # En-têtes
        headers = ["EMPLOIS ET RESSOURCES", "Montant", "Pourcentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_start, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Données
        data = [
            ("EMPLOIS STABLES", float(bilan.emplois_stables), ""),
            ("Ressources stables", float(bilan.ressources_stables), ""),
            ("FRNG", float(bilan.frng), "=B11-B10"),
            ("", "", ""),
            ("ACTIFS CIRCULANTS", float(bilan.actifs_circulants), ""),
            ("Passifs circulants", float(bilan.passifs_circulants), ""),
            ("BFR", float(bilan.bfr), "=B14-B13"),
            ("", "", ""),
            ("TRÉSORERIE ACTIVE", float(bilan.tresorerie_active), ""),
            ("Trésorerie passive", float(bilan.tresorerie_passive), ""),
            ("TRÉSORERIE NETTE", float(bilan.tresorerie_nette), "=B18-B17"),
        ]
        
        for i, (label, montant, formule) in enumerate(data, row_start + 1):
            ws.cell(row=i, column=1, value=label)
            
            if formule:
                ws.cell(row=i, column=2, value=formule)
            else:
                ws.cell(row=i, column=2, value=montant)
            
            ws.cell(row=i, column=3, value="")
            
            # Style
            for col in range(1, 4):
                cell = ws.cell(row=i, column=col)
                cell.border = self.thin_border
                
                if i in [row_start + 3, row_start + 7, row_start + 11]:  # Sous-totaux
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                elif i == row_start + 12:  # Total
                    cell.font = self.total_font
                    cell.fill = self.total_fill
                else:
                    cell.alignment = Alignment(horizontal='left' if col == 1 else 'right')
        
        # Formatage des nombres
        for row in range(row_start + 1, row_start + 13):
            cell = ws.cell(row=row, column=2)
            cell.number_format = '#,##0.00'
        
        # Largeur des colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12

    def create_bilan_fonctionnel_analysis(self, ws: Worksheet, bilan: BilanFonctionnel, options: Dict[str, Any]):
        """Créer la feuille d'analyse du bilan fonctionnel."""
        # Titre
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "ANALYSE FONCTIONNELLE"
        title_cell.font = self.title_font
        title_cell.fill = self.title_fill
        title_cell.alignment = self.title_alignment
        
        # Tableau d'analyse
        row_start = 3
        
        # En-têtes
        headers = ["Indicateur", "Valeur", "Seuil", "Interprétation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_start, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Données d'analyse
        frng = float(bilan.frng)
        bfr = float(bilan.bfr)
        treso_nette = float(bilan.tresorerie_nette)
        
        data = [
            ("FRNG", frng, "> 0", "Positif" if frng > 0 else "Négatif"),
            ("BFR", bfr, "Variable", "Positif" if bfr > 0 else "Négatif"),
            ("Trésorerie nette", treso_nette, "> 0", "Positive" if treso_nette > 0 else "Négative"),
            ("Équilibre", abs(frng - (bfr + treso_nette)), "= 0", "Équilibré" if abs(frng - (bfr + treso_nette)) < 0.01 else "Déséquilibré"),
        ]
        
        for i, (indicateur, valeur, seuil, interpretation) in enumerate(data, row_start + 1):
            ws.cell(row=i, column=1, value=indicateur)
            ws.cell(row=i, column=2, value=valeur)
            ws.cell(row=i, column=3, value=seuil)
            ws.cell(row=i, column=4, value=interpretation)
            
            # Style
            for col in range(1, 5):
                cell = ws.cell(row=i, column=col)
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center')
                
                if interpretation in ["Positif", "Positive", "Équilibré"]:
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif interpretation in ["Négatif", "Négative", "Déséquilibré"]:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Formatage des nombres
        for row in range(row_start + 1, row_start + 5):
            cell = ws.cell(row=row, column=2)
            cell.number_format = '#,##0.00'
        
        # Largeur des colonnes
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 15

    def create_bilan_fonctionnel_charts(self, ws: Worksheet, bilan: BilanFonctionnel, options: Dict[str, Any]):
        """Créer les graphiques du bilan fonctionnel."""
        # Données pour les graphiques
        chart_data = [
            ["Rubrique", "Montant"],
            ["Emplois stables", float(bilan.emplois_stables)],
            ["Ressources stables", float(bilan.ressources_stables)],
            ["FRNG", float(bilan.frng)],
            ["Actifs circulants", float(bilan.actifs_circulants)],
            ["Passifs circulants", float(bilan.passifs_circulants)],
            ["BFR", float(bilan.bfr)],
            ["Trésorerie active", float(bilan.tresorerie_active)],
            ["Trésorerie passive", float(bilan.tresorerie_passive)],
            ["Trésorerie nette", float(bilan.tresorerie_nette)],
        ]
        
        # Ajouter les données
        for row_idx, row_data in enumerate(chart_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Graphique en barres
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Structure financière"
        chart.y_axis.title = "Montant"
        chart.x_axis.title = "Rubriques"
        
        # Références des données
        data_ref = Reference(ws, min_col=2, min_row=2, max_row=len(chart_data), max_col=2)
        labels_ref = Reference(ws, min_col=1, min_row=2, max_row=len(chart_data), max_col=1)
        
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(labels_ref)
        
        # Position du graphique
        ws.add_chart(chart, "F2")

    def create_bilan_financier_sheets(self, wb: Workbook, bilan: BilanFinancier, options: Dict[str, Any]):
        """Créer les feuilles pour le bilan financier."""
        # Feuille Actif
        ws_actif = wb.create_sheet("Actif")
        self.create_actif_sheet(ws_actif, bilan, options)
        
        # Feuille Passif
        ws_passif = wb.create_sheet("Passif")
        self.create_passif_sheet(ws_passif, bilan, options)
        
        # Feuille Ratios
        ws_ratios = wb.create_sheet("Ratios")
        self.create_ratios_sheet(ws_ratios, bilan, options)

    def create_actif_sheet(self, ws: Worksheet, bilan: BilanFinancier, options: Dict[str, Any]):
        """Créer la feuille de l'actif."""
        # Titre
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = "ACTIF"
        title_cell.font = self.title_font
        title_cell.fill = self.title_fill
        title_cell.alignment = self.title_alignment
        
        # Tableau
        row_start = 3
        
        # En-têtes
        headers = ["Rubriques", "Montant", "Pourcentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_start, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Données
        total_actif = float(bilan.total_actif)
        data = [
            ("Immobilisations nettes", float(bilan.immobilisations_nettes)),
            ("Stocks", float(bilan.stocks)),
            ("Créances clients", float(bilan.creances_clients)),
            ("Autres créances", float(bilan.autres_creances)),
            ("Trésorerie active", float(bilan.tresorerie_active)),
            ("TOTAL ACTIF", total_actif),
        ]
        
        for i, (label, montant) in enumerate(data, row_start + 1):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=montant)
            
            # Pourcentage
            if i == row_start + 5:  # Total
                ws.cell(row=i, column=3, value="100%")
            else:
                pct = (montant / total_actif * 100) if total_actif > 0 else 0
                ws.cell(row=i, column=3, value=f"=B{i}/B{row_start + 5}")
            
            # Style
            for col in range(1, 4):
                cell = ws.cell(row=i, column=col)
                cell.border = self.thin_border
                
                if i == row_start + 5:  # Total
                    cell.font = self.total_font
                    cell.fill = self.total_fill
                else:
                    cell.alignment = Alignment(horizontal='left' if col == 1 else 'right')
        
        # Formatage
        for row in range(row_start + 1, row_start + 6):
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3).number_format = '0.00%'

    def create_passif_sheet(self, ws: Worksheet, bilan: BilanFinancier, options: Dict[str, Any]):
        """Créer la feuille du passif."""
        # Titre
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = "PASSIF"
        title_cell.font = self.title_font
        title_cell.fill = self.title_fill
        title_cell.alignment = self.title_alignment
        
        # Tableau
        row_start = 3
        
        # En-têtes
        headers = ["Rubriques", "Montant", "Pourcentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_start, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Données
        total_passif = float(bilan.total_passif)
        data = [
            ("Capital social", float(bilan.capital_social)),
            ("Réserves", float(bilan.reserves)),
            ("Résultat net", float(bilan.resultat_net)),
            ("Capitaux propres", float(bilan.capitaux_propres)),
            ("Dettes financières LT", float(bilan.dettes_financieres_lt)),
            ("Dettes fournisseurs", float(bilan.dettes_fournisseurs)),
            ("Autres dettes CT", float(bilan.autres_dettes_ct)),
            ("Trésorerie passive", float(bilan.tresorerie_passive)),
            ("TOTAL PASSIF", total_passif),
        ]
        
        for i, (label, montant) in enumerate(data, row_start + 1):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=montant)
            
            # Pourcentage
            if i == row_start + 8:  # Total
                ws.cell(row=i, column=3, value="100%")
            else:
                pct = (montant / total_passif * 100) if total_passif > 0 else 0
                ws.cell(row=i, column=3, value=f"=B{i}/B{row_start + 8}")
            
            # Style
            for col in range(1, 4):
                cell = ws.cell(row=i, column=col)
                cell.border = self.thin_border
                
                if i == row_start + 4:  # Sous-total capitaux propres
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                elif i == row_start + 8:  # Total
                    cell.font = self.total_font
                    cell.fill = self.total_fill
                else:
                    cell.alignment = Alignment(horizontal='left' if col == 1 else 'right')
        
        # Formatage
        for row in range(row_start + 1, row_start + 9):
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3).number_format = '0.00%'

    def create_ratios_sheet(self, ws: Worksheet, bilan: BilanFinancier, options: Dict[str, Any]):
        """Créer la feuille des ratios."""
        # Titre
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "RATIOS FINANCIERS"
        title_cell.font = self.title_font
        title_cell.fill = self.title_fill
        title_cell.alignment = self.title_alignment
        
        # Calcul des ratios
        total_actif = float(bilan.total_actif)
        total_passif = float(bilan.total_passif)
        capitaux_propres = float(bilan.capitaux_propres)
        dettes_totales = total_passif - capitaux_propres
        
        ratios = [
            ("Ratio d'endettement", dettes_totales / total_actif if total_actif > 0 else 0, "< 70%"),
            ("Ratio d'autonomie", capitaux_propres / total_passif if total_passif > 0 else 0, "> 50%"),
            ("Ratio de liquidité générale", (float(bilan.stocks) + float(bilan.creances_clients) + float(bilan.tresorerie_active)) / float(bilan.dettes_fournisseurs) if float(bilan.dettes_fournisseurs) > 0 else 0, "> 1"),
        ]
        
        # Tableau
        row_start = 3
        
        # En-têtes
        headers = ["Ratio", "Valeur", "Formule", "Recommandation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_start, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Données
        for i, (ratio, valeur, recommandation) in enumerate(ratios, row_start + 1):
            ws.cell(row=i, column=1, value=ratio)
            ws.cell(row=i, column=2, value=valeur)
            ws.cell(row=i, column=3, value=self._get_ratio_formula(ratio))
            ws.cell(row=i, column=4, value=recommandation)
            
            # Style
            for col in range(1, 5):
                cell = ws.cell(row=i, column=col)
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center')
                
                # Coloration selon la valeur
                if col == 2:
                    if ratio == "Ratio d'endettement" and valeur < 0.7:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif ratio == "Ratio d'autonomie" and valeur > 0.5:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif ratio == "Ratio de liquidité générale" and valeur > 1:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        
        # Formatage
        for row in range(row_start + 1, row_start + 4):
            ws.cell(row=row, column=2).number_format = '0.00%'

    def create_patrimoine_sheets(self, wb: Workbook, patrimoine: PatrimoineEntreprise, options: Dict[str, Any]):
        """Créer les feuilles pour le patrimoine."""
        # Feuille principale
        ws = wb.create_sheet("Patrimoine")
        self.create_patrimoine_main(ws, patrimoine, options)
        
        # Feuille d'analyse
        ws_analysis = wb.create_sheet("Analyse Patrimoniale")
        self.create_patrimoine_analysis(ws_analysis, patrimoine, options)

    def create_patrimoine_main(self, ws: Worksheet, patrimoine: PatrimoineEntreprise, options: Dict[str, Any]):
        """Créer la feuille principale du patrimoine."""
        # Titre
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = "PATRIMOINE DE L'ENTREPRISE"
        title_cell.font = self.title_font
        title_cell.fill = self.title_fill
        title_cell.alignment = self.title_alignment
        
        # Tableau
        row_start = 3
        
        # En-têtes
        headers = ["Éléments patrimoniaux", "Montant", "Pourcentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_start, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Données
        patrimoine_net = float(patrimoine.patrimoine_net)
        data = [
            ("Actifs économiques", float(patrimoine.actifs_economiques)),
            ("Dettes financières", float(patrimoine.dettes_financieres)),
            ("Actif net comptable", float(patrimoine.actif_net_comptable)),
            ("Capitaux propres retraités", float(patrimoine.capitaux_propres_retraites)),
            ("PATRIMOINE NET", patrimoine_net),
        ]
        
        for i, (label, montant) in enumerate(data, row_start + 1):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=montant)
            
            # Pourcentage
            if i == row_start + 4:  # Total
                ws.cell(row=i, column=3, value="100%")
            else:
                pct = (montant / patrimoine_net * 100) if patrimoine_net > 0 else 0
                ws.cell(row=i, column=3, value=f"=B{i}/B{row_start + 4}")
            
            # Style
            for col in range(1, 4):
                cell = ws.cell(row=i, column=col)
                cell.border = self.thin_border
                
                if i == row_start + 4:  # Total
                    cell.font = self.total_font
                    cell.fill = self.total_fill
                else:
                    cell.alignment = Alignment(horizontal='left' if col == 1 else 'right')
        
        # Formatage
        for row in range(row_start + 1, row_start + 5):
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=3).number_format = '0.00%'

    def create_patrimoine_analysis(self, ws: Worksheet, patrimoine: PatrimoineEntreprise, options: Dict[str, Any]):
        """Créer la feuille d'analyse patrimoniale."""
        # Titre
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "ANALYSE PATRIMONIALE"
        title_cell.font = self.title_font
        title_cell.fill = self.title_fill
        title_cell.alignment = self.title_alignment
        
        # Tableau des ratios
        row_start = 3
        
        # En-têtes
        headers = ["Ratio", "Valeur", "Seuil", "Interprétation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_start, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Données
        data = [
            ("Ratio d'endettement", patrimoine.ratio_endettement or 0, "< 50%", self._interpret_ratio(patrimoine.ratio_endettement, 0.5, 0.8)),
            ("Ratio de solvabilité", patrimoine.ratio_solvabilite or 0, "> 1", self._interpret_solvability(patrimoine.ratio_solvabilite)),
            ("Ratio de liquidité", patrimoine.ratio_liquidite or 0, "> 1", self._interpret_ratio(patrimoine.ratio_liquidite, 1.0, 0.8)),
        ]
        
        for i, (ratio, valeur, seuil, interpretation) in enumerate(data, row_start + 1):
            ws.cell(row=i, column=1, value=ratio)
            ws.cell(row=i, column=2, value=valeur)
            ws.cell(row=i, column=3, value=seuil)
            ws.cell(row=i, column=4, value=interpretation)
            
            # Style
            for col in range(1, 5):
                cell = ws.cell(row=i, column=col)
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center')
                
                # Coloration selon l'interprétation
                if interpretation.startswith("✅"):
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif interpretation.startswith("⚠️"):
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Formatage
        for row in range(row_start + 1, row_start + 4):
            ws.cell(row=row, column=2).number_format = '0.00%'

    def create_summary_sheet(self, wb: Workbook, report_data, options: Dict[str, Any]):
        """Créer la feuille de résumé."""
        ws = wb.create_sheet("Résumé")
        ws.move_sheet("Résumé", 0)  # Déplacer en première position
        
        # Titre
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "RÉSUMÉ DU RAPPORT"
        title_cell.font = self.title_font
        title_cell.fill = self.title_fill
        title_cell.alignment = self.title_alignment
        
        # Informations générales
        ws['A3'] = "Entreprise:"
        ws['B3'] = options.get('entreprise', 'Entreprise')
        ws['A4'] = "Type de rapport:"
        
        if isinstance(report_data, BilanFonctionnel):
            ws['B4'] = "Bilan fonctionnel"
        elif isinstance(report_data, BilanFinancier):
            ws['B4'] = "Bilan financier"
        elif isinstance(report_data, PatrimoineEntreprise):
            ws['B4'] = "Patrimoine de l'entreprise"
        
        ws['A5'] = "Période:"
        ws['B5'] = options.get('periode', '2024')
        ws['A6'] = "Devise:"
        ws['B6'] = options.get('devise', 'MAD')
        ws['A7'] = "Date de génération:"
        ws['B7'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        # Points clés
        ws.merge_cells('A9:D9')
        subtitle_cell = ws['A9']
        subtitle_cell.value = "POINTS CLÉS"
        subtitle_cell.font = self.subtitle_font
        subtitle_cell.fill = self.subtitle_fill
        subtitle_cell.alignment = self.title_alignment
        
        # Contenu des points clés selon le type de rapport
        row_start = 11
        
        if isinstance(report_data, BilanFonctionnel):
            points_cles = [
                ("FRNG", float(report_data.frng), "Positif" if float(report_data.frng) > 0 else "Négatif"),
                ("BFR", float(report_data.bfr), "Positif" if float(report_data.bfr) > 0 else "Négatif"),
                ("Trésorerie nette", float(report_data.tresorerie_nette), "Positive" if float(report_data.tresorerie_nette) > 0 else "Négative"),
            ]
        elif isinstance(report_data, BilanFinancier):
            total_actif = float(report_data.total_actif)
            points_cles = [
                ("Total actif", total_actif, ""),
                ("Capitaux propres", float(report_data.capitaux_propres), f"{float(report_data.capitaux_propres)/total_actif*100:.1f}%" if total_actif > 0 else ""),
                ("Endettement", total_actif - float(report_data.capitaux_propres), f"{(total_actif - float(report_data.capitaux_propres))/total_actif*100:.1f}%" if total_actif > 0 else ""),
            ]
        else:  # PatrimoineEntreprise
            points_cles = [
                ("Patrimoine net", float(report_data.patrimoine_net), ""),
                ("Ratio d'endettement", f"{report_data.ratio_endettement or 0:.1%}", ""),
                ("Ratio de solvabilité", f"{report_data.ratio_solvabilite or 0:.2f}", ""),
            ]
        
        for i, (label, valeur, complement) in enumerate(points_cles, row_start):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=valeur)
            ws.cell(row=i, column=3, value=complement)
            
            # Style
            for col in range(1, 4):
                cell = ws.cell(row=i, column=col)
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='left' if col == 1 else 'right')
        
        # Largeur des colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    # Méthodes utilitaires
    def _get_ratio_formula(self, ratio_name: str) -> str:
        """Obtenir la formule pour un ratio."""
        formulas = {
            "Ratio d'endettement": "Dettes / Actif total",
            "Ratio d'autonomie": "Capitaux propres / Passif total",
            "Ratio de liquidité générale": "(Stocks + Créances + Trésorerie) / Dettes CT",
        }
        return formulas.get(ratio_name, "")

    def _interpret_ratio(self, ratio: Optional[float], good_threshold: float, bad_threshold: float) -> str:
        """Interpréter un ratio."""
        if ratio is None:
            return "Non calculable"
        
        if ratio <= good_threshold:
            return "✅ Bon"
        elif ratio >= bad_threshold:
            return "⚠️ À surveiller"
        else:
            return "🟡 Moyen"

    def _interpret_solvability(self, ratio: Optional[float]) -> str:
        """Interpréter le ratio de solvabilité."""
        if ratio is None:
            return "Non calculable"
        
        if ratio > 1:
            return "✅ Solvable"
        elif ratio > 0.5:
            return "🟡 Solvabilité moyenne"
        else:
            return "⚠️ Solvabilité faible"
